"""
ETL script for the Skilled Migration Pathway Dashboard.

Reads the Department of Home Affairs "Australian Migration Statistics 2024-25"
statistical package (source: data.gov.au) and produces clean, tidy CSV files
ready for the Streamlit dashboard.

Source file: raw_data/migration_trends_statistical_package_2024_25.xlsx
Output folder: data/
"""

import csv
import os
import openpyxl

RAW_FILE = os.path.join(os.path.dirname(__file__), "raw_data", "migration_trends_statistical_package_2024_25.xlsx")
OUT_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(OUT_DIR, exist_ok=True)

wb = openpyxl.load_workbook(RAW_FILE, data_only=True)


def write_csv(filename, header, rows):
    path = os.path.join(OUT_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    print(f"Wrote {path} ({len(rows)} rows)")


def clean_year(y):
    if y is None:
        return None
    return str(y).replace("–", "-").strip()


def is_number(v):
    return isinstance(v, (int, float))


# ---------------------------------------------------------------------------
# Table 1.0: Australia's Migration Program outcome, 1984-85 to 2024-25
# ---------------------------------------------------------------------------
ws = wb["1.0"]
rows = []
for row in ws.iter_rows(min_row=4, values_only=True):
    year = clean_year(row[1])
    if not year or not is_number(row[2]):
        continue
    skill, family, child, special, total = row[2], row[3], row[4], row[5], row[6]
    rows.append([year, skill, family, child, special, total])
write_csv("1_migration_program_outcome_by_year.csv",
          ["year", "skill_stream", "family_stream", "child_stream", "special_eligibility", "total"], rows)

# ---------------------------------------------------------------------------
# Table 1.1: Migration Program outcome by category, 2015-16 to 2024-25
# long format: year, category, count
# ---------------------------------------------------------------------------
ws = wb["1.1"]
categories = [
    "Employer Sponsored", "State/Territory Nominated", "Regional", "Skilled Independent",
    "Global Talent (Independent)", "Business Innovation & Investment", "Distinguished Talent",
    "National Innovation", "Skilled Regional", "Skill stream total",
    "Partner", "Parent", "Child (family)", "Other Family", "Family stream total",
    "Child stream total", "Special Eligibility total", "Total",
]
rows = []
for row in ws.iter_rows(min_row=4, values_only=True):
    year = clean_year(row[1])
    if not year:
        continue
    values = row[2:2 + len(categories)]
    for cat, val in zip(categories, values):
        if is_number(val):
            rows.append([year, cat, val])
write_csv("2_migration_program_by_category.csv", ["year", "category", "count"], rows)

# ---------------------------------------------------------------------------
# Table 1.3 / 1.6 / 1.8: top citizenship countries by pathway
# ---------------------------------------------------------------------------
def extract_country_table(sheet_name, pathway, min_row=4):
    ws = wb[sheet_name]
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    countries = [c for c in header_row[2:] if c and str(c).strip().lower() != "total"]
    rows_out = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        year = clean_year(row[1])
        if not year:
            continue
        values = row[2:2 + len(countries)]
        for country, val in zip(countries, values):
            if is_number(val):
                rows_out.append([year, pathway, str(country).rstrip("0123456789"), val])
    return rows_out


all_countries = []
all_countries += extract_country_table("1.3", "Employer Sponsored")
all_countries += extract_country_table("1.6", "Regional")
all_countries += extract_country_table("1.8", "State/Territory Nominated")
write_csv("3_top_citizenship_countries_by_pathway.csv", ["year", "pathway", "country", "count"], all_countries)

# ---------------------------------------------------------------------------
# Table 1.5 / 1.7 / 1.9: top nominated occupations by pathway (ANZSCO)
# ---------------------------------------------------------------------------
def extract_occupation_table(sheet_name, pathway):
    ws = wb[sheet_name]
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    occs = [c for c in header_row[2:] if c and str(c).strip().lower() != "total"]
    rows_out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        year = clean_year(row[1])
        if not year:
            continue
        values = row[2:2 + len(occs)]
        for occ, val in zip(occs, values):
            if is_number(val):
                rows_out.append([year, pathway, str(occ).rstrip("0123456789").strip(), val])
    return rows_out


all_occupations = []
all_occupations += extract_occupation_table("1.5", "Employer Sponsored")
all_occupations += extract_occupation_table("1.7", "Regional")
all_occupations += extract_occupation_table("1.9", "State/Territory Nominated")
write_csv("4_top_occupations_by_pathway.csv", ["year", "pathway", "occupation", "count"], all_occupations)

# ---------------------------------------------------------------------------
# Table 2.0: Temporary visas granted by visa category, 2001-02 to 2024-25
# ---------------------------------------------------------------------------
ws = wb["2.0"]
header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
tv_categories = [c for c in header_row[2:] if c and str(c).strip().lower() != "total"]
rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    year = clean_year(row[1])
    if not year:
        continue
    values = row[2:2 + len(tv_categories)]
    for cat, val in zip(tv_categories, values):
        if is_number(val):
            rows.append([year, str(cat).rstrip("012"), val])
write_csv("5_temporary_visas_by_category.csv", ["year", "category", "count"], rows)

print("\nETL complete.")
